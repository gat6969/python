import openpyxl


n = int(input("Enter the number of people: "))


wb = openpyxl.Workbook()
sheet = wb.active


sheet["A1"] = "Name"
sheet["B1"] = "Age"
sheet["C1"] = "Place"


for i in range(2, n + 2):
    name = input(f"Enter name of person {i-1}: ")
    age = input(f"Enter age of person {i-1}: ")
    place = input(f"Enter place of person {i-1}: ")
    
    sheet[f"A{i}"] = name
    sheet[f"B{i}"] = age
    sheet[f"C{i}"] = place


wb.save("people_data.xlsx")


wb = openpyxl.load_workbook("people_data.xlsx")
sheet = wb.active

print("\nData from the Excel file:")
for row in sheet.iter_rows(min_row=2, values_only=True):
    print(f"Name: {row[0]}, Age: {row[1]}, Place: {row[2]}")
